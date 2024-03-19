import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from openpyxl import load_workbook
from utility.utility import *
from tkinter import Toplevel, Label, Button, N, S, E, W
#NOTE:
# modified to staff-roles, not reliant on staff names anymore




def clockinclockout(frame):
	
	with open(find_file('staff-names.txt'), "r", encoding = "utf-8") as file:
		names = (file.read()).split("\n")
	

	# Load or create the workbook
	log_location = find_file('staff-log.xlsx')
	remove_blank_rows(log_location)
	
	

	# Create the main window and set its size
	input_var = tk.StringVar()
	
	message = tk.Message(
		frame,
		text ="Names",
		width=500,
	)
	# Create a listbox for the names and set the font size
	name_listbox = tk.Listbox(
		frame, listvariable=tk.Variable(value=names), width=50, height=20, font = ('Helvetica', 15)
	)
	name_listbox.pack(pady=20)



	# Function to handle clocking in
	def clock_in():
		workbook = load_workbook(filename=log_location)
		# Check if a name is selected
		if not name_listbox.curselection():
			messagebox.showinfo("Clock In", "Names not chosen")
			
			return
		# Get the selected name
		name = name_listbox.get(name_listbox.curselection())

		# print(name)
		now = datetime.now()
		date = now.strftime("%d/%m/%Y")
		for i, row in enumerate(
			workbook.active.iter_rows(min_row=2, values_only=True), start=2
		):
			# print(row)
			if row[0] == date and row[1] == name:
				if row[3] == "" or row[3] == None:
					messagebox.showinfo(
						"Clock In", f"{name} already logged in"
					)
					return 


		if now.minute < 10:
			time = f"{now.hour}:0{now.minute}"
		else:
			time = f"{now.hour}:{now.minute}"

		messagebox.showinfo("Clock In", f"{name} logged in at {date} {time}")

		workbook.active.append([date, name, time, ""])
		workbook.save(filename=log_location)
		workbook.close()

	def clock_out():
		workbook = load_workbook(filename=log_location)
		input_str = input_var.get()


		if not name_listbox.curselection():
			messagebox.showinfo("Clock Out", "No names chosen")
			return
		if not input_str:
			messagebox.showinfo(
				"Clock Out", "Missing Patient Input"
			)
			return


		try:
			int(input_str)
		except:
			messagebox.showinfo("Clock Out", "Not a number")
			return

		# Get the selected name
		name = name_listbox.get(name_listbox.curselection())


		now = datetime.now()
		date = now.strftime("%d/%m/%Y")
		if now.minute < 10:
			time = f"{now.hour}:0{now.minute}"
		else:
			time = f"{now.hour}:{now.minute}"

		for i, row in enumerate(workbook.active.iter_rows(min_row=2, values_only=True), start=2):
			# print(row[0])
			# print(type(row[0]))
			if (
				now.strftime("%d/%m/%Y") == row[0]
				and row[1] == name
				and row[2]
				and not row[3]
			):
				# Update the clock out time

				workbook.active.cell(row=i, column=4, value=time)
				workbook.active.cell(row=i, column=5, value=int(input_str))
				workbook.save(filename=log_location)
				workbook.close()
				# Show a message box
				messagebox.showinfo("Clock Out", f"{name} logged out at {date} {time}")

				return
		workbook.close()
		# If the user has not clocked in, show a message box

	clockinout = tk.Frame(frame)
	clockinout.pack()

	# Create the clock in button
	clock_in_button = tk.Button(
		clockinout, text="Clock In", command=clock_in, width=8, height=1,font=('Helvetica', 15)
	)
	clock_in_button.pack(pady = 10, side = tk.LEFT, padx = 20)

	# Create the clock out button
	clock_out_button = tk.Button(
		clockinout, text="Clock Out", command=clock_out, width= 8, height=1,font=('Helvetica', 15)
	)
	clock_out_button.pack(pady=10, side = tk.LEFT)

	message = tk.Message(
		frame,
		text="Input the number of patients:",
		width=500,
		justify="center",
		font=('Helvetica', 15)
	)
	message.pack(pady=10)

	entry = tk.Entry(frame, textvariable=input_var, width=50, font=('Helvetica', 15))
	entry.pack(pady=20)


	message = tk.Message(
		frame,
		text="To edit the name list, go to settings and choose 'Staff Names'",
		width=500,
		font=('Helvetica', 15)
	)
	message.pack()

	# Start the main loop
	

if __name__ == '__main__':
	ws = tk.Tk()

	clockinclockout(ws)
	ws.mainloop()